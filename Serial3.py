import serial
import serial.tools.list_ports
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import threading
import requests
import json
import datetime
import configparser
import uuid
import sqlite3
import openpyxl
from pathlib import Path
from tkinter import filedialog



class SerialInterface:
    def __init__(self, root):
        self.root = root
        self.mostrar_ventana_inicio_sesion()
        
        self.root.title("Comunicación WebService Deprisa")
        root.iconbitmap('montra.ico')
        
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True)
        self.medicion_tab = ttk.Frame(self.notebook)
        self.configuracion_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.medicion_tab, text="Medición", state="disabled")  # Inicialmente deshabilitada
        self.notebook.add(self.configuracion_tab, text="Configuración", state="disabled")  # Inicialmente deshabilitada
        #self.notebook.add(self.medicion_tab, text="Medición") #Se comenta para prueba de inicio de sesión
        #self.notebook.add(self.configuracion_tab, text="Configuración") #Se comenta para prueba de inicio de sesión
        
        self.create_medicion_tab()
        self.create_configuracion_tab()
        self.cargar_configuracion()
        self.root.protocol("WM_DELETE_WINDOW", self.cerrar_aplicacion)

        self.paquetes_enviados = 0
        self.paquetes_no_enviados = 0
        
        self.tiempo_espera = 2  # Tiempo en segundos para esperar la recepción de datos
        self.datos_recibidos = False  # Agrega esta línea para inicializar la variable

        #self.notebook.bind("<<NotebookTabChanged>>", self.verificar_contraseña)

        self.verificar_mac()
        
        
        #print(self.get_mac_address())

#VERIFICACIÓN DE MAC
    #Obtener mac_address
    def get_mac_address(self):
        mac = uuid.UUID(int=uuid.getnode()).hex[-12:]
        formatted_mac = ':'.join([mac[i:i+2] for i in range(0, 12, 2)])
        return formatted_mac

    #Verificar mac con la lista
    def verificar_mac(self):
        direcciones_mac_permitidas = ["30:05:05:b8:bb:31", "bc:f1:71:f3:5f:60", "30:05:05:b8:bb:35"]  # Lista de direcciones MAC permitidas  # Reemplaza con la MAC permitida
        mac_actual = self.get_mac_address()  # Usa el método de obtener_mac() definido
        if mac_actual not in direcciones_mac_permitidas:
            self.cerrar_puerto()
            self.mostrar_error_y_salir()  # Llama a la función para mostrar un mensaje de error y cerrar el programa

    
        # Agrega esta función para mostrar un mensaje de error y cerrar el programa.
    
    #Si no es autorizada, muestra error y cierra el programa.
    def mostrar_error_y_salir(self):
        mensaje = "Este software solo puede ejecutarse en una computadora autorizada."
        messagebox.showerror("Error", mensaje)
        root.destroy()  # Cierra la aplicación

#CREACIÓN DE VENTANA DE INICIO DE SESIÓN
    def mostrar_ventana_inicio_sesion(self):
        self.root.withdraw()  # Oculta la ventana principal
        self.ventana_inicio_sesion = tk.Toplevel(self.root)
        self.ventana_inicio_sesion.title("Inicio de Sesión")
        self.ventana_inicio_sesion.iconbitmap('montra.ico')

        usuario_label = tk.Label(self.ventana_inicio_sesion, text="Usuario:")
        usuario_label.pack()
        self.usuario_entry = tk.Entry(self.ventana_inicio_sesion)
        self.usuario_entry.pack()


        contrasena_label = tk.Label(self.ventana_inicio_sesion, text="Contraseña:")
        contrasena_label.pack()
        self.contrasena_entry = tk.Entry(self.ventana_inicio_sesion, show="*")  # Muestra asteriscos para ocultar la contraseña
        self.contrasena_entry.pack()

        # Botón de inicio de sesión
        boton_login = tk.Button(self.ventana_inicio_sesion, text="Iniciar Sesión", command=self.verificar_credenciales)
        boton_login.pack()
        
        self.contrasena_entry.bind('<Return>', lambda event=None: self.verificar_credenciales())

        
        self.ventana_inicio_sesion.protocol("WM_DELETE_WINDOW", self.cerrar_aplicacion)
        

    def verificar_credenciales(self):
        usuario = self.usuario_entry.get()
        contrasena = self.contrasena_entry.get()

        conn = sqlite3.connect('Montradb.db')
        cursor = conn.cursor()

        cursor.execute("SELECT Acceso FROM Login WHERE Usuario=? AND Contraseña=?", (usuario, contrasena))
        resultado = cursor.fetchone()

        if resultado:
            acceso = resultado[0]
            if acceso == 'ADMINISTRADOR':
                self.notebook.tab(0, state="normal")  # Habilitar la pestaña de Medición
                self.notebook.tab(1, state="normal")  # Habilitar la pestaña de Configuración
                self.notebook.select(0)  # Cambiar a la pestaña de Medición
                self.ventana_inicio_sesion.destroy()  # Cerrar la ventana de inicio de sesión
            elif acceso == 'OPERARIO':
                self.notebook.tab(0, state="normal")  # Habilitar la pestaña de Medición
                self.notebook.select(0)  # Cambiar a la pestaña de Medición
                self.ventana_inicio_sesion.destroy()  # Cerrar la ventana de inicio de sesión
            self.root.deiconify()  # Mostrar la ventana principal nuevamente
        else:
            messagebox.showerror("Error", "Credenciales incorrectas. Intente nuevamente.")
            # Borra el contenido de los campos de entrada
            #self.mostrar_ventana_inicio_sesion()
            self.usuario_entry.delete(0, tk.END)
            self.contrasena_entry.delete(0, tk.END)
            
        conn.close()
        self.usuario_registrado=usuario

#CREACIÓN DE VENTANA DE MEDICIÓN
    def create_medicion_tab(self):
        
        self.sku_var = tk.StringVar()
        self.length_var = tk.StringVar()
        self.width_var = tk.StringVar()
        self.height_var = tk.StringVar()
        self.weight_var = tk.StringVar()
        self.response_text = tk.StringVar()


        ttk.Label(self.medicion_tab, text="SKU:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.sku_entry = ttk.Entry(self.medicion_tab, textvariable=self.sku_var, font=('Helvetica', 10), width=22)
        self.sku_entry.grid(row=0, column=1, padx=10, pady=5, ipadx=15)
        self.sku_entry.focus_set()
        self.medir_button = ttk.Button(self.medicion_tab, text="Medir", command=self.enviar_trama)
        self.medir_button.grid(row=1, columnspan=2, padx=10, pady=5)
        
        self.send_button = ttk.Button(self.medicion_tab, text="Enviar", command=self.send_data)
        self.send_button.grid(row=2, columnspan=2, padx=10, pady=5)

        ttk.Label(self.medicion_tab, text="Largo:").grid(row=0, column=2, padx=10, pady=5, sticky="w")
        self.largo_entry = ttk.Entry(self.medicion_tab, textvariable=self.length_var)
        self.largo_entry.grid(row=0, column=3, padx=10, pady=5)

        ttk.Label(self.medicion_tab, text="Ancho:").grid(row=1, column=2, padx=10, pady=5, sticky="w")
        self.ancho_entry = ttk.Entry(self.medicion_tab, textvariable=self.width_var)
        self.ancho_entry.grid(row=1, column=3, padx=10, pady=5)

        ttk.Label(self.medicion_tab, text="Alto:").grid(row=2, column=2, padx=10, pady=5, sticky="w")
        self.alto_entry = ttk.Entry(self.medicion_tab, textvariable=self.height_var)
        self.alto_entry.grid(row=2, column=3, padx=10, pady=5)

        ttk.Label(self.medicion_tab, text="Peso:").grid(row=3, column=2, padx=10, pady=5, sticky="w")
        self.peso_entry = ttk.Entry(self.medicion_tab, textvariable=self.weight_var)
        self.peso_entry.grid(row=3, column=3, padx=10, pady=5)

        ttk.Label(self.medicion_tab, text="Respuesta:").grid(row=8, column=0, columnspan=2, padx=5, pady=10, sticky="w")
        self.response_entry = ttk.Entry(self.medicion_tab, textvariable=self.response_text, state="readonly")
        self.response_entry.grid(row=9, column=0, columnspan=3, pady=5, sticky="nsew")
        
        # Crear la tabla para mostrar los datos
        columns = ('Sku', 'Largo', 'Ancho', 'Alto', 'Peso', 'Fecha', 'Usuario')
        self.tree = ttk.Treeview(self.medicion_tab, columns=columns, show='headings')

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column('Sku', width=200)
            self.tree.column('Largo', width=50)
            self.tree.column('Ancho', width=50)
            self.tree.column('Alto', width=50)
            self.tree.column('Peso', width=50)
            self.tree.column('Fecha', width=130)
            self.tree.column('Usuario', width=80)
            #self.tree.column(col, width=100)

        self.tree.grid(row=10, column=0, columnspan=15)
        
        # Aplicar un estilo con bordes a la tabla
        style = ttk.Style()
        style.configure("Treeview", font=('Helvetica', 9), rowheight=20)
        style.configure("Treeview.Heading", font=('Helvetica', 9))
        style.configure("Treeview.Treeview", borderwidth=1)  # Esto añade bordes alrededor de cada celda
        
        # Crear barras de desplazamiento
        y_scroll = ttk.Scrollbar(self.medicion_tab, orient="vertical", command=self.tree.yview)
        y_scroll.grid(row=10, column=16, sticky='ns')
        self.tree.configure(yscrollcommand=y_scroll.set)

        x_scroll = ttk.Scrollbar(self.medicion_tab, orient="horizontal", command=self.tree.xview)
        x_scroll.grid(row = 11, column=0, columnspan=16, sticky = tk.W+tk.E)
        self.tree.configure(xscrollcommand=x_scroll.set)

        #Etiquetas del contador
        self.paquetes_enviados_label = tk.Label(self.medicion_tab, text="Envíos exitosos: 0", font=("verdama", 10))
        self.paquetes_enviados_label.grid(row=3, column=0, columnspan=2)

        self.paquetes_no_enviados_label = tk.Label(self.medicion_tab, text="Envíos fallidos: 0", font=("Verdana", 10))
        self.paquetes_no_enviados_label.grid(row=4,column=0, columnspan=2)
        
        self.medir_button.bind('<Return>', self.on_enter_press)
        self.medir_button.bind('<FocusIn>', self.on_button_focus_in)
        self.medir_button.bind('<FocusOut>', self.on_button_focus_out)
        
        self.send_button.bind('<Return>', self.on_enter_press)
        self.send_button.bind('<FocusIn>', self.on_sendbutton_focus_in)
        self.send_button.bind('<FocusOut>', self.on_sendbutton_focus_out)
        
        self.sku_entry.bind("<Return>", self.cambiar_foco_a_medir)
        
        self.cerrar_sesion_button = ttk.Button(self.medicion_tab, text="Cerrar Sesión", command=self.cerrar_sesion)
        self.cerrar_sesion_button.grid(row=5, column=0, columnspan=2, padx=10, pady=5)

        
        #exportar_button = tk.Button(self.medicion_tab, text="Exportar a Excel", command=self.exportar_excel)
        #exportar_button.grid(row=12, column=0, columnspan=2)

    #Función para funcion del boton de cerrar sesión
    def cerrar_sesion(self):
        self.notebook.tab(0, state="disabled")  # Deshabilitar la pestaña de Medición
        self.notebook.tab(1, state="disabled")  # Deshabilitar la pestaña de Configuración
        self.mostrar_ventana_inicio_sesion()  # Mostrar la ventana de inicio de sesión nuevamente

    #CREACIÓN DE COMANDOS PARA FOCUS Y ACCIONES CON ENTER
    # Evento de ENTER en SKU
    def cambiar_foco_a_medir(self, event): 
        self.medir_button.focus_set()
    
    #Evento de enter al tener focus en "Medir" o "Enviar"
    def on_enter_press(self, event):
        if self.is_medirbutton_focused:
            self.enviar_trama()
        elif self.is_sendbutton_focused:
            self.send_data()
    
    #Confirmar cursor en boton "medir" con variable en TRUE
    def on_button_focus_in(self, event): 
        self.is_medirbutton_focused = True 
    
    #Confirmar cursor en boton "medir" con variable en FALSE
    def on_button_focus_out(self, event): 
        self.is_medirbutton_focused = False    
    
    #Confirmar cursor en boton "Enviar" con variable en TRUE
    def on_sendbutton_focus_in(self, event):
        self.is_sendbutton_focused = True
    
    #Confirmar cursor en boton "Enviar" con variable en TRUE
    def on_sendbutton_focus_out(self, event):
        self.is_sendbutton_focused = False


#CREACIÓN DE VENTANA DE CONFIGURACIÓN
    def create_configuracion_tab(self):

        self.url_var = tk.StringVar()
        self.username_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.machine_name_var = tk.StringVar()
        self.ruta_exportacion = tk.StringVar()
        

        ttk.Label(self.configuracion_tab, text="URL del Web Service:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        url_entry = ttk.Entry(self.configuracion_tab, textvariable=self.url_var, show="*")
        url_entry.grid(row=0, column=1, padx=10, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="Usuario:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        username_entry = ttk.Entry(self.configuracion_tab, textvariable=self.username_var, show="*")
        username_entry.grid(row=1, column=1, padx=10, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="Contraseña:").grid(row=2, column=0, padx=10, pady=5, sticky="w")
        password_entry = ttk.Entry(self.configuracion_tab, textvariable=self.password_var, show="*")
        password_entry.grid(row=2, column=1, padx=10, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="Nombre de la Máquina:").grid(row=3, column=0, padx=10, pady=5, sticky="w")
        machine_name_entry = ttk.Entry(self.configuracion_tab, textvariable=self.machine_name_var)
        machine_name_entry.grid(row=3, column=1, padx=10, pady=5, sticky="w")
        
        ttk.Label(self.configuracion_tab, text="Ruta exportación").grid(row=4, column=0, padx=10, pady=5, sticky="w")
        ruta_exportacion_entry = ttk.Entry(self.configuracion_tab, textvariable=self.ruta_exportacion)
        ruta_exportacion_entry.grid(row=4, column=1, padx=10, pady=5, sticky="w")
        
        self.actualizar_puertos_button = ttk.Button(self.configuracion_tab, text="Carpeta", command=self.seleccionar_carpeta)
        self.actualizar_puertos_button.grid(row=4, column=2, padx=10, pady=5)
                
        ttk.Label(self.configuracion_tab, text="Puertos COM disponibles:").grid(row=5, column=0, padx=10, pady=5, sticky="w")
        self.puertos_combobox = ttk.Combobox(self.configuracion_tab)
        self.puertos_combobox.grid(row=5, column=1, padx=10, pady=5)

        self.actualizar_puertos_button = ttk.Button(self.configuracion_tab, text="Actualizar", command=self.listar_puertos)
        self.actualizar_puertos_button.grid(row=5, column=2, padx=10, pady=5)

        self.abrir_puerto_button = ttk.Button(self.configuracion_tab, text="Abrir Puerto", command=self.abrir_puerto)
        self.abrir_puerto_button.grid(row=6, column=1, padx=10, pady=5)

        self.cerrar_puerto_button = ttk.Button(self.configuracion_tab, text="Cerrar Puerto", command=self.cerrar_puerto)
        self.cerrar_puerto_button.grid(row=7, column=1, padx=10, pady=5)
        self.cerrar_puerto_button.configure(state="disabled")
        
        self.guardar_config_button = ttk.Button(self.configuracion_tab, text="Guardar Configuración", command=self.guardar_configuracion)
        self.guardar_config_button.grid(row=8, columnspan=2, padx=10, pady=5)

        #Botón para crear usuarios
        crear_usuario_button = ttk.Button(self.configuracion_tab, text="Crear usuario", command=self.abrir_ventana_crear_usuario)
        crear_usuario_button.grid(row=9, columnspan=2, padx=10, pady=5)

    #Configuración de boton para escoger carpeta de exportación
    def seleccionar_carpeta(self):
        folder_selected = filedialog.askdirectory(title="Seleccione una carpeta de destino")
        self.ruta_exportacion.set(folder_selected)

#CONFIGURACIÓN DE PUERTOS
    #Listar los puertos disponibles
    def listar_puertos(self):
        puertos = [puerto.device for puerto in serial.tools.list_ports.comports()]
        self.puertos_combobox['values'] = puertos
    
    #Abrir puerto seleccionado
    def abrir_puerto(self): 
        puerto_seleccionado = self.puertos_combobox.get()
        try:
            self.puerto_serial = serial.Serial(puerto_seleccionado, baudrate=9600)
            self.puertos_combobox.configure(state="disabled")
            self.abrir_puerto_button.configure(state="disabled")
            self.cerrar_puerto_button.configure(state="enabled")

            # Guardar el último puerto en el archivo config.ini
            self.guardar_configuracion()
            
            self.data_thread = threading.Thread(target=self.leer_datos)
            self.data_thread.start()
        except Exception as e:
            mensaje = f"Error al abrir el puerto: {e}"
            messagebox.showerror("Error", mensaje)
    
    #Cerrar el puerto
    def cerrar_puerto(self): 
        if hasattr(self, 'puerto_serial') and self.puerto_serial.is_open:
            self.puerto_serial.close()
            self.puerto_serial = None
            self.puertos_combobox.configure(state="readonly")  # Desbloquear el combobox
            self.abrir_puerto_button.configure(state="enabled")
            self.cerrar_puerto_button.configure(state="disabled")

#CONFIGURACIÓN DE ARCHIVO.INI PARA PRECARGAR Y GUARDAR LOS DATOS
    #Precargar configuración en archivo.ini
    def cargar_configuracion(self):
        config = configparser.ConfigParser()
        config.read('config.ini')
        if 'Configuracion' in config:
            self.url_var.set(config['Configuracion'].get('url', ''))
            self.username_var.set(config['Configuracion'].get('username', ''))
            self.password_var.set(config['Configuracion'].get('password', ''))
            self.machine_name_var.set(config['Configuracion'].get('machine_name', ''))
            self.ruta_exportacion.set(config['Configuracion'].get('ruta_exportacion', ''))
            # Cargar y establecer el último puerto en el combobox
            ultimo_puerto = config['Configuracion'].get('ultimo_puerto', '')
            self.puertos_combobox.set(ultimo_puerto)
            self.contraseña_actual = config['Configuracion'].get('contraseña_adicional', 'MONTRA101') # Obtener la contraseña
            self.abrir_puerto()

    #Guardar configuración en archivo.ini
    def guardar_configuracion(self):
        config = configparser.ConfigParser()
        config.read('config.ini')
        # Guardar los valores de configuración
        config['Configuracion']['url'] = self.url_var.get()
        config['Configuracion']['username'] = self.username_var.get()
        config['Configuracion']['password'] = self.password_var.get()
        config['Configuracion']['machine_name'] = self.machine_name_var.get()
        config['Configuracion']['ruta_exportacion'] = self.ruta_exportacion.get()
        # Obtener el último puerto seleccionado del combobox
        ultimo_puerto = self.puertos_combobox.get()
        config['Configuracion']['ultimo_puerto'] = ultimo_puerto
        config['Configuracion']['contraseña_adicional'] = self.contraseña_actual
        with open('config.ini', 'w') as configfile:
            config.write(configfile)

#CONFIGURACIÓN PARA EXPORTACIÓN DE DATOS
    def exportar_excel(self):
        self.ruta_destino = Path(self.ruta_exportacion.get())
        fecha_actual = datetime.datetime.now().strftime("%Y%d%m_%H-%M-%S")
        nombre_archivo = f"CubiScan_{fecha_actual}.xlsx"
        ruta_completa = self.ruta_destino / nombre_archivo  # Usar pathlib para construir la ruta

    # Verificar si la carpeta de destino existe
        if not self.ruta_destino.exists() or not self.ruta_destino.is_dir():
            if folder_selected := filedialog.askdirectory(
                title="Selecciona la carpeta de destino"
            ):
                self.ruta_exportacion.set(folder_selected)
                self.nueva_ruta_destino = Path(self.ruta_exportacion.get())  # Actualizar la ruta de destino
                ruta_completa = self.nueva_ruta_destino / nombre_archivo  # Usar pathlib para construir la ruta
            else:
                folder_selected = filedialog.askdirectory(title="Seleccione una carpeta de destino")
                self.ruta_destino=Path(self.ruta_exportacion.get())
                ruta_completa = self.ruta_destino / nombre_archivo  # Usar pathlib para construir la ruta
                return
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Medidas"
        
        # Encabezados
        encabezados = ["SKU", "Largo", "Ancho", "Alto", "Peso", "Fecha"]
        for col_num, encabezado in enumerate(encabezados, 1):
            worksheet.cell(row=1, column=col_num, value=encabezado)

        # Datos
        for row_num, item in enumerate(self.tree.get_children(), 2):
            datos_fila = [self.tree.item(item, 'values')[0], self.tree.item(item, 'values')[1],
                        self.tree.item(item, 'values')[2], self.tree.item(item, 'values')[3],
                        self.tree.item(item, 'values')[4], self.tree.item(item, 'values')[5]]
            for col_num, valor in enumerate(datos_fila, 1):
                worksheet.cell(row=row_num, column=col_num, value=valor)

        # Guardar el archivo Excel
        #print(self.ruta_exportacion.get())
        self.guardar_configuracion()
        workbook.save(ruta_completa)

#CONFIGURACIÓN DE CONTRASEÑA PARA ACCESO A VENTANA CONFIGURACIÓN
    #Verificar contraseña guardada con contraseña ingresada al cambiar de pestaña. Contraseña inicial: MONTRA101
    """def verificar_contraseña(self, event):
        if self.notebook.tab(self.notebook.select(), "text") == "Configuración":
            password = simpledialog.askstring("Contraseña", "Ingrese la contraseña:", show="*")
            if password != self.contraseña_actual:  # Usar la contraseña actual almacenada
                self.notebook.select(self.medicion_tab)
                messagebox.showerror("Acceso denegado", "La contraseña ingresada es incorrecta. Acceso denegado a la pestaña de configuración.")"""

    #Creación de ventana para cambiar contraseña.
    def abrir_ventana_crear_usuario(self):
        # Ventana emergente para cambiar la contraseña
        self.crear_usuario_window = tk.Toplevel(self.root)
        self.crear_usuario_window.title("Cambiar Contraseña")
        self.crear_usuario_window.iconbitmap('montra.ico')
        self.crear_usuario_window.grab_set()


        ttk.Label(self.crear_usuario_window, text="Usuario:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        nombre_usuario_entry = ttk.Entry(self.crear_usuario_window)
        nombre_usuario_entry.grid(row=0, column=1, padx=10, pady=5)

        ttk.Label(self.crear_usuario_window, text="Contraseña:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        contraseña_usuario_entry = ttk.Entry(self.crear_usuario_window, show="*")
        contraseña_usuario_entry.grid(row=1, column=1, padx=10, pady=5)

        ttk.Label(self.crear_usuario_window, text="Perfil:").grid(row=2, column=0, padx=10, pady=5, sticky="w")
        acceso_combobox = ttk.Combobox(self.crear_usuario_window, state="readonly", values=["ADMINISTRADOR", "OPERARIO"])
        acceso_combobox.grid(row=2, column=1, padx=10, pady=5)

        ttk.Button(self.crear_usuario_window, text="Guardar",command=lambda: self.guardar_nuevo_usuario(nombre_usuario_entry.get(), contraseña_usuario_entry.get(),acceso_combobox.get() )).grid(row=4, columnspan=2, padx=10, pady=5)
    
    #Acción ejecutada por el boton para guardar la nueva contraseña en el archivo.ini
    def guardar_nuevo_usuario(self, nombre_usuario_entry, contraseña_usuario_entry, acceso_combobox):
        
        if (nombre_usuario_entry!="" and contraseña_usuario_entry!="" and acceso_combobox!=""):
            conn = sqlite3.connect('Montradb.db')
            cursor = conn.cursor()
            cursor.execute('INSERT INTO Login (Usuario, Contraseña, Acceso) VALUES (?, ?, ?)', (nombre_usuario_entry, contraseña_usuario_entry, acceso_combobox ))
            conn.commit()
            conn.close()
            messagebox.showinfo(message="El usuario se ha creado con exito")
            self.crear_usuario_window.destroy()  # Cerrar la ventana de inicio de sesión
        else:
            messagebox.showerror(message="Debe completar todos los campos para crear usuario")


    #Configuración para cerrar el puerto y guardar la configuración al cerrar la app.
    def cerrar_aplicacion(self):
        if hasattr(self, 'puerto_serial') and self.puerto_serial and self.puerto_serial.is_open:
            self.cerrar_puerto()  # Cerrar el puerto si está abierto
        self.guardar_configuracion()  # Guardar la configuración antes de salir
        self.exportar_excel()
        self.root.destroy()  # Cerrar la aplicación


#CONFIGURACIÓN PARA RECEPCIÓN PEDIR Y RECIBIR DATOS DE CUBISCAN
    #Solicitar medición
    def enviar_trama(self):
        self.iniciar_espera_datos()
        if hasattr(self, 'puerto_serial') and self.puerto_serial.is_open:
            #self.iniciar_espera_datos()  # Iniciar la espera de datos
            trama = b'\x02M\x03\r\n'  # Trama: <STX>M<ETX><CR><LF>
            enviar_thread = threading.Thread(target=self.enviar_trama_thread, args=(trama,))
            enviar_thread.start()
    
    #Metodo independiente para solicitar medición
    def enviar_trama_thread(self, trama):
        try:
            self.puerto_serial.write(trama)
        except Exception as e:
            print("Error al enviar la trama:", e)  

    #Recepción de datos y estructurar en campos:
    def leer_datos(self):
        while hasattr(self, 'puerto_serial') and self.puerto_serial and self.puerto_serial.is_open:
            try:
                dato = self.puerto_serial.readline().decode('utf-8')
                if dato:
                    self.datos_recibidos = True  # Datos recibidos, detener temporizador
                    if "\x02" in dato and "\x03" in dato:
                        trama = dato.split("\x02")[1].split("\x03")[0]
                        valores = trama.split(",")                        
                        for valor in valores:
                            if valor.startswith("L"):
                                largo = valor.split("L")[1]
                                largo = round(float(largo))  # Convertir a número, redondear
                                self.largo_entry.delete(0, tk.END)
                                self.largo_entry.insert(0, str(largo))
                            elif valor.startswith("W"):
                                ancho = valor.split("W")[1]
                                ancho = round(float(ancho))  # Convertir a número, redondear
                                self.ancho_entry.delete(0, tk.END)
                                self.ancho_entry.insert(0, str(ancho))
                            elif valor.startswith("H"):
                                alto = valor.split("H")[1]
                                alto = round(float(alto))  # Convertir a número, redondear
                                self.alto_entry.delete(0, tk.END)
                                self.alto_entry.insert(0, str(alto))
                            elif valor.startswith("K"):
                                peso = valor.split("K")[1]
                                peso = float(peso)
                                self.peso_entry.delete(0, tk.END)
                                self.peso_entry.insert(0, peso)
                    self.verificar_datos_cubiscan()
            except:
                pass
    
    #Confirmación que ningun dato recibido sea cero.
    def verificar_datos_cubiscan(self):
        # Verificar si alguno de los campos está en 0
        if self.sku_var.get() <= '0' or self.length_var.get() <= '0' or self.width_var.get() <= '0' or self.weight_var.get() <= '0' and self.length_var.get()!= "" or self.width_var.get() == "" or self.height_var.get() == "" or self.weight_var.get() == "":
            self.medir_button.focus_set()
            messagebox.showerror("Error", "Los campos SKU, Largo, Ancho y Alto no pueden ser 0 o estar vacíos.")
            return  # No se envía la información si algún campo es 0
        else:
            self.send_button.focus_set()

    # Metodo para iniciar espera de datos
    def iniciar_espera_datos(self):
        self.datos_recibidos = False
        self.timer = threading.Timer(self.tiempo_espera, self.mostrar_error_sin_datos)
        self.timer.start()

    # Método para mostrar un mensaje de error si no se reciben datos
    def mostrar_error_sin_datos(self):
        if not self.datos_recibidos:
            messagebox.showerror("Error", "No hay comunicación con la CubiScan, revise cables y conexiones.")

    # Agrega este método para detener el temporizador
    def detener_espera_datos(self):
        if hasattr(self, 'timer') or self.timer.is_alive():
            self.timer.cancel()
    

#CONFIGURACIÓN ENVÍO JSON
    #Configuración del envío de estructura JSON
    def send_data(self):
        # Obtener los valores de los campos
        sku = self.sku_var.get()
        largo = self.length_var.get()
        ancho = self.width_var.get()
        alto = self.height_var.get()
        peso = self.weight_var.get()
        fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        Usuario = self.usuario_registrado

        # Guardar datos en la base de datos
        conn = sqlite3.connect('Montradb.db')
        cursor = conn.cursor()
        cursor.execute('INSERT INTO Montra (sku, largo, ancho, alto, peso, fecha) VALUES (?, ?, ?, ?,?, ?)', (sku, largo, ancho, alto, peso, fecha))
        conn.commit()
        conn.close()
        
        # Construir el JSON con los datos ingresados
        data = {
            "machine_pid": self.machine_name_var.get(),
            "code": self.sku_var.get(),
            "measure_date": fecha,
            "length": self.length_var.get(),
            "width": self.width_var.get(),
            "heigth": self.height_var.get(),
            "weigth": self.weight_var.get(),
            "unit_type": "cm"
        }
        print(data)
        
        # Verificar si alguno de los campos está en 0
        if sku <= '0' or largo <= '0' or ancho <= '0' or alto <= '0' or peso <= '0':
            self.send_button.focus_set()
            messagebox.showerror("Error", "Los campos SKU, Largo, Ancho y Alto no pueden ser 0 o estar vacíos.")
            return  # No se envía la información si algún campo es 0
        elif sku == "" or largo == "" or alto == "" or ancho == "" or peso=="":
            self.send_button.focus_set()
        else:
            # Mostrar datos en la tabla
            self.tree.insert('', 'end', values=(sku, largo, ancho, alto, peso, fecha, Usuario))


        # Realizar la solicitud POST al WebService
        url = self.url_var.get()
        headers = {"Content-Type": "application/json"}
        response = requests.post(url, data=json.dumps(data), headers=headers, auth=(self.username_var.get(), self.password_var.get()))
        
        #contador
        if response.status_code == 200:
            self.paquetes_enviados += 1
            
        else:
            self.paquetes_no_enviados += 1
        # Actualizar la respuesta en la interfaz
        self.response_text.set(response.text)
        #actualizar contadores
        self.update_contadores()
        self.sku_var.set("")     # Borra el contenido del campo SKU
        self.length_var.set("")  # Borra el contenido del campo Largo
        self.width_var.set("")   # Borra el contenido del campo Ancho
        self.height_var.set("")  # Borra el contenido del campo Alto
        self.weight_var.set("")  # Borra el contenido del campo Peso


    #Conteos exitoso y fallidos de envío
    def update_contadores(self):
        self.paquetes_enviados_label.config(text=f"Envíos exitosos: {self.paquetes_enviados}")
        self.paquetes_no_enviados_label.config(text=f"Envíos fallidos: {self.paquetes_no_enviados}")


if __name__ == "__main__":
    root = tk.Tk()
    root.resizable(False,False)
    app = SerialInterface(root)
    root.mainloop()
