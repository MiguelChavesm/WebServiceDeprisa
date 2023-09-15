import serial
import serial.tools.list_ports
import tkinter as tk
from tkinter import ttk, messagebox
import threading
import requests
import json
import datetime
import configparser
import os
import time
import uuid
import sqlite3
import openpyxl
from pathlib import Path
from tkinter import filedialog
import customtkinter
from PIL import Image


class SerialInterface:
    def __init__(self, root):
        self.root = root
        self.root.title("MONTRA")
        root.iconbitmap('Icons/montra.ico')
        # Define el ancho y alto de la ventana

        self.direcciones_mac_permitidas = ["4C-44-5B-95-52-85", "bc:f1:71:f3:5f:60", "30-05-05-B8-BB-35"]  # Lista de direcciones MAC permitidas  # Reemplaza con la MAC permitida
        self.texto_licencia="Desarrollado por Grupo Montra\nUso exclusivo para Deprisa\n\nLicencia: Deprisa Cartagena"
        self.perfil_acceso= ""
        self.valores_combox=[]
        self.mostrar_ventana_inicio_sesion()
        
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True)
        self.medicion_tab = ttk.Frame(self.notebook)
        self.configuracion_tab = ttk.Frame(self.notebook)
        self.usuarios_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.medicion_tab, text="Medición", state="disabled")  # Inicialmente deshabilitada
        self.notebook.add(self.usuarios_tab, text="Usuarios", state="disabled")  # Inicialmente deshabilitada)
        self.notebook.add(self.configuracion_tab, text="Configuración", state="disabled")  # Inicialmente deshabilitada
        self.create_medicion_tab()
        self.create_configuracion_tab()
        self.abrir_ventana_crear_usuario()
        self.cargar_configuracion()
        
        self.root.protocol("WM_DELETE_WINDOW", self.cerrar_aplicacion)
        self.fecha_limite = (2023, 10, 15, 13, 45)
        
        self.verificar_fecha_limite_periodicamente()
        
        self.paquetes_enviados = 0
        self.paquetes_no_enviados = 0
        self.tiempo_espera = 2  # Tiempo en segundos para esperar la recepción de datos
        self.datos_recibidos = False  # Agrega esta línea para inicializar la variable

        #print(self.get_mac_address())


#CREACION LICENCIA TEMPORAL
    def verificar_fecha_limite(self):
        # Obtiene la fecha y hora actual
        fecha_actual = time.localtime()

        # Compara la fecha actual con la fecha límite
        if fecha_actual >= self.fecha_limite:
            # Si ha pasado la fecha límite, muestra un mensaje y cierra la aplicación
            mensaje = "La versión de prueba ha expirado. No puedes usar la aplicación."
            messagebox.showerror("Versión de Prueba Expirada", mensaje)
            self.cerrar_aplicacion()
    
    def verificar_fecha_limite_periodicamente(self):
            # Programa la verificación de la fecha límite cada segundo
            self.verificar_fecha_limite()
            self.root.after(1000, self.verificar_fecha_limite_periodicamente)



#VERIFICACIÓN DE MAC
    #Obtener mac_address
    def get_mac_address(self):
        mac = uuid.UUID(int=uuid.getnode()).hex[-12:]
        formatted_mac = ':'.join([mac[i:i+2] for i in range(0, 12, 2)])
        formatted_mac=formatted_mac.upper()
        formatted_mac = formatted_mac.replace(":", "-")
        return formatted_mac

#CREACIÓN DE VENTANA DE INICIO DE SESIÓN
    def mostrar_ventana_inicio_sesion(self):
        self.root.withdraw()  # Oculta la ventana principal
        self.ventana_inicio_sesion = tk.Toplevel(self.root)
        self.ventana_inicio_sesion.title("MONTRA")
        self.ventana_inicio_sesion.iconbitmap('Icons/montra.ico')
        self.ventana_inicio_sesion.configure(bg="#FFFFFF")
        self.ventana_inicio_sesion.resizable(False,False)
        Fuente_inicio_sesion = ("Helvetica", 13)
        # Cargar imagen del disco.
        self.logo_montra = tk.PhotoImage(file="Icons/Logo_Montra.png")
        self.logo_montra = self.logo_montra.subsample(1, 1)
        # Insertarla en una etiqueta.
        self.label_logo_montra = ttk.Label(self.ventana_inicio_sesion, image=self.logo_montra, background="#FFFFFF")
        self.label_logo_montra.pack(padx=(60,60), pady=(20,0))

        self.logo_cubiscan = tk.PhotoImage(file="Icons/Cubiscan_logo.png")
        self.logo_cubiscan = self.logo_cubiscan.subsample(1, 1)
        # Insertarla en una etiqueta.
        self.label_logo_cubiscan = ttk.Label(self.ventana_inicio_sesion, image=self.logo_cubiscan, background="#FFFFFF")
        self.label_logo_cubiscan.pack()

        
        usuario_label = tk.Label(self.ventana_inicio_sesion, text="Usuario:", background="#FFFFFF", font=Fuente_inicio_sesion)
        usuario_label.pack(padx=(50,0),pady=(20, 0), anchor="w")  # Agregar espacio arriba
        self.usuario_entry = tk.Entry(self.ventana_inicio_sesion, borderwidth=2, font=Fuente_inicio_sesion, width=30, background="#E4E9EA")
        self.usuario_entry.pack(padx=(50,50))


        contrasena_label = tk.Label(self.ventana_inicio_sesion, text="Contraseña:", background="#FFFFFF", font=Fuente_inicio_sesion)
        contrasena_label.pack(padx=(50,0),pady=(20, 0), anchor="w")
        self.contrasena_entry = tk.Entry(self.ventana_inicio_sesion, show="*", borderwidth=2, font=Fuente_inicio_sesion, width=30, background="#E4E9EA")  # Muestra asteriscos para ocultar la contraseña
        self.contrasena_entry.pack(padx=(50,50))

        # Botón de inicio de sesión
        login_image = customtkinter.CTkImage(Image.open("Icons/login2.png").resize((100,100), Image.Resampling.LANCZOS))
        boton_login = customtkinter.CTkButton(self.ventana_inicio_sesion, text="Iniciar Sesión", font=("Helvetica", 16), text_color="#000000", fg_color="#B1B3B6", hover_color="#828890", width=200, height=20, compound="left", image= login_image, command=self.verificar_credenciales)
        boton_login.pack(pady=(20, 0))
        self.contrasena_entry.bind('<Return>', lambda event=None: self.verificar_credenciales())
        
        self.logo_deprisa0 = tk.PhotoImage(file="Icons/Deprisa_logo.png")
        self.logo_deprisa0 = self.logo_deprisa0.subsample(1, 1)
        # Insertarla en una etiqueta.
        self.label_logo_deprisa = ttk.Label(self.ventana_inicio_sesion, image=self.logo_deprisa0, background="#FFFFFF")
        self.label_logo_deprisa.pack(pady=(50, 10))
        
        
        self.ventana_inicio_sesion.protocol("WM_DELETE_WINDOW", self.cerrar_aplicacion)
        

    def verificar_credenciales(self):
        usuario = self.usuario_entry.get()
        contrasena = self.contrasena_entry.get()

        conn = sqlite3.connect('Montradb.db')
        cursor = conn.cursor()

        cursor.execute("SELECT Acceso FROM Login WHERE Usuario=? AND Contraseña=?", (usuario, contrasena))
        resultado = cursor.fetchone()

        if resultado:
            acceso = resultado[0]
            if acceso == 'ADMINISTRADOR':
                self.notebook.tab(0, state="normal")  # Habilitar la pestaña de Medición
                self.notebook.tab(1, state="normal")  # Habilitar la pestaña de Usuarios
                self.notebook.select(0)  # Cambiar a la pestaña de Medición
                self.acceso_combobox['values'] = [""]
                self.acceso_combobox['values']=["ADMINISTRADOR","OPERARIO"]
                self.ventana_inicio_sesion.destroy()  # Cerrar la ventana de inicio de sesión
            elif acceso == 'OPERARIO':
                self.notebook.tab(0, state="normal")  # Habilitar la pestaña de Medición
                self.notebook.select(0)  # Cambiar a la pestaña de Medición
                self.perfil_acceso=2
                self.ventana_inicio_sesion.destroy()  # Cerrar la ventana de inicio de sesión
            elif acceso == 'SUPERUSUARIO':
                self.notebook.tab(0, state="normal")  # Habilitar la pestaña de Medición
                self.notebook.tab(1, state="normal")  # Habilitar la pestaña de Usuarios
                self.notebook.tab(2, state="normal")  # Habilitar la pestaña de Configuración
                self.notebook.select(0)  # Cambiar a la pestaña de Medición
                self.acceso_combobox['values'] = [""]
                self.acceso_combobox['values']=["ADMINISTRADOR","OPERARIO", "SUPERUSUARIO"]
                self.ventana_inicio_sesion.destroy()  # Cerrar la ventana de inicio de sesión
            self.root.deiconify() # Mostrar la ventana principal nuevamente
            self.sku_entry.focus_set()
        else:
            messagebox.showerror("Error", "Credenciales incorrectas. Intente nuevamente.")
            # Borra el contenido de los campos de entrada
            self.usuario_entry.delete(0, tk.END)
            self.contrasena_entry.delete(0, tk.END)
            
        conn.close()
        self.usuario_registrado=usuario
        self.perfil_acceso=acceso

#CREACIÓN VENTANA DE USUARIOS
    def abrir_ventana_crear_usuario(self):
        # Ventana emergente para cambiar la contraseña
        self.colorbackground= "lightgrey"
        self.background = ttk.Label(self.usuarios_tab, background=self.colorbackground)
        self.background.grid(row=0, column=0, rowspan=9,padx=(0,20), sticky="snew")
        
        self.label_montra4 = ttk.Label(self.usuarios_tab, image=self.logo_montra2, background=self.colorbackground)
        self.label_montra4.grid(row=1, column=0, rowspan=3, padx=(15,25), pady=(10,0), sticky="s")

        self.label_cubiscan4 = ttk.Label(self.usuarios_tab, image=self.logo_cubiscan2,background=self.colorbackground)
        self.label_cubiscan4.grid(row=4, column=0, rowspan=3, padx=(15,25), sticky="n")
        
        self.label_deprisa2 = ttk.Label(self.usuarios_tab, image=self.logo_deprisa, background=self.colorbackground)
        self.label_deprisa2.grid(row=5, column=0, rowspan=1, padx=(15,25), pady=10, sticky="ew")
        
        logout_image = customtkinter.CTkImage(Image.open("Icons/logout.png").resize((100,100), Image.Resampling.LANCZOS))
        boton_logout = customtkinter.CTkButton(self.usuarios_tab, text="Cerrar Sesión", corner_radius=1,font=("Helvetica", 14), text_color="#000000", fg_color="#FFFFFF", hover_color="#828890", width=150, height=20, compound="left", image= logout_image, command=self.cerrar_sesion)
        boton_logout.grid(row=6, column=0, columnspan=1, padx=(10,30), pady=5, sticky="new")
        
        ttk.Label(self.usuarios_tab, text=self.texto_licencia ,background=self.colorbackground, font=("Arial", 9)).grid(row=7, rowspan=2, column=0,pady=(20,30), padx=(5,0), sticky="w")

        ttk.Label(self.usuarios_tab, text="Usuario:").grid(row=1, column=2, padx=(50,0), pady=5, sticky="w")
        self.nombre_usuario_entry = ttk.Entry(self.usuarios_tab)
        self.nombre_usuario_entry.grid(row=1, column=3, padx=(0,0), pady=5)

        ttk.Label(self.usuarios_tab, text="Contraseña:").grid(row=2, column=2, padx=(50,0), pady=5, sticky="w")
        self.contraseña_usuario_entry = ttk.Entry(self.usuarios_tab, show="*")
        self.contraseña_usuario_entry.grid(row=2, column=3, padx=(0,0), pady=5)

        ttk.Label(self.usuarios_tab, text="Perfil:").grid(row=3, column=2, padx=(50,0), pady=5, sticky="w")
        
        self.acceso_combobox = ttk.Combobox(self.usuarios_tab, state="readonly")
        self.acceso_combobox.grid(row=3, column=3, padx=(0,0), pady=5)

        ttk.Button(self.usuarios_tab, text="Guardar",command=lambda: self.guardar_nuevo_usuario()).grid(row=4, column=3, columnspan=2, padx=10, pady=5)

        
        # Crear la tabla para mostrar los datos
        columns = ('Usuario', 'Contraseña', 'Perfil')
        self.tree2 = ttk.Treeview(self.usuarios_tab, columns=columns, show='headings')

        for col in columns:
            self.tree2.heading(col, text=col)
            self.tree2.column('Usuario', width=200)
            self.tree2.column('Contraseña', width=200)
            self.tree2.column('Perfil', width=200)

        self.tree2.grid(row=5, column=1, rowspan=1, columnspan=10, pady=(10,10))
        
        self.actualizar_tabla()
        
    def actualizar_tabla (self):
        self.tree2.tag_configure('ADMINISTRADOR', background='#FF6666') #ROJO
        self.tree2.tag_configure('OPERARIO', background='#B7FF66') #VERDE
        self.tree2.tag_configure('SUPERUSUARIO', background='#66DCFF') #AZUL
        self.tree2.yview_moveto(1.0)  # Desplaza la vista hacia el final de la tabla
        conn = sqlite3.connect('Montradb.db')
        cursor = conn.cursor()
        # Limpiar la tabla actual
        for item in self.tree2.get_children():
            self.tree2.delete(item)
        # Consultar la base de datos y agregar los datos al Treeview
        cursor.execute("SELECT * FROM LogIn")
        for row in cursor.fetchall():
            access_type = row[2] # Obtener el tipo de acceso de la fila
            contraseña_oculta = self.ocultar_contraseña(row[1])
            row_with_asterisks = (row[0], contraseña_oculta, row[2])
            self.tree2.insert("", "end", values=row_with_asterisks, tags=(access_type.upper(),))
        conn.close()
        
    def ocultar_contraseña(self, contraseña):
        return '*' * len(contraseña)

    #Acción ejecutada por el boton para guardar la nueva contraseña en el archivo.ini
    def guardar_nuevo_usuario(self):
        nombre_usuario_entry = self.nombre_usuario_entry.get()
        contraseña_usuario_entry = self.contraseña_usuario_entry.get()
        acceso_combobox = self.acceso_combobox.get()
        
        if (nombre_usuario_entry!="" and contraseña_usuario_entry!="" and acceso_combobox!=""):
            conn = sqlite3.connect('Montradb.db')
            cursor = conn.cursor()
            cursor.execute('INSERT INTO Login (Usuario, Contraseña, Acceso) VALUES (?, ?, ?)', (nombre_usuario_entry, contraseña_usuario_entry, acceso_combobox ))
            conn.commit()
            messagebox.showinfo(message="El usuario se ha creado con exito")
            self.actualizar_tabla()
            conn.close()

            self.nombre_usuario_entry.delete(0, 'end')
            self.contraseña_usuario_entry.delete(0, 'end')
            self.acceso_combobox.set('')  # Borra la selección del Combobox
        else:
            messagebox.showerror(message="Debe completar todos los campos para crear usuario")


#CREACION DE METODO DE CIERRE DE APLICACIÓN.
    def cerrar_aplicacion(self):
        if hasattr(self, 'puerto_serial') and self.puerto_serial and self.puerto_serial.is_open:
            self.cerrar_puerto()  # Cerrar el puerto si está abierto
        self.guardar_configuracion()  # Guardar la configuración antes de salir
        self.exportar_excel()
        self.exportar_log()
        self.root.destroy()  # Cerrar la aplicación

#CREACIÓN DE VENTANA DE MEDICIÓN
    def create_medicion_tab(self):
        self.sku_var = tk.StringVar()
        self.length_var = tk.StringVar()
        self.width_var = tk.StringVar()
        self.height_var = tk.StringVar()
        self.weight_var = tk.StringVar()
        self.response_text = tk.StringVar()

        # Insertarla en una etiqueta.
        self.colorbackground= "lightgrey"
        self.background = ttk.Label(self.medicion_tab, background=self.colorbackground)
        self.background.grid(row=0, column=0, rowspan=9,padx=(0,20), sticky="snew")
        
        self.logo_montra2 = tk.PhotoImage(file="Icons/Logo_Montra3.png")
        self.logo_montra2 = self.logo_montra2.subsample(1, 1)
        self.label_montra2 = ttk.Label(self.medicion_tab, image=self.logo_montra2, background=self.colorbackground)
        self.label_montra2.grid(row=0, column=0, rowspan=3, padx=(10,20), pady=(10,0), sticky="s")
        
        self.logo_deprisa = tk.PhotoImage(file="Icons/Deprisa_logo.png")
        self.logo_deprisa = self.logo_deprisa.subsample(1, 1)
        self.label_deprisa1 = ttk.Label(self.medicion_tab, image=self.logo_deprisa, background=self.colorbackground)
        self.label_deprisa1.grid(row=4, column=0, rowspan=2, padx=(15,20), pady=10, sticky="ew")
        
        self.logo_cubiscan2 = tk.PhotoImage(file="Icons/Cubiscan_logo.png")
        self.logo_cubiscan2 = self.logo_cubiscan2.subsample(1, 1)
        self.label_cubiscan2 = ttk.Label(self.medicion_tab, image=self.logo_cubiscan2,background=self.colorbackground)
        self.label_cubiscan2.grid(row=3, column=0, rowspan=3, padx=(5,20), sticky="n")

        # Botón de cerrar de sesión
        logout_image = customtkinter.CTkImage(Image.open("Icons/logout.png").resize((100,100), Image.Resampling.LANCZOS))
        boton_logout = customtkinter.CTkButton(self.medicion_tab, text="Cerrar Sesión", corner_radius=1,font=("Helvetica", 14), text_color="#000000", fg_color="#FFFFFF", hover_color="#828890", width=200, height=20, compound="left", image= logout_image, command=self.cerrar_sesion)
        boton_logout.grid(row=5, column=0, columnspan=1, padx=(10,30), pady=5, sticky="new")

        ttk.Label(self.medicion_tab, text=self.texto_licencia ,background=self.colorbackground, font=("Arial", 9)).grid(row=6, rowspan=1, column=0, padx=(5,0), pady=(0,5), sticky="w")
        
        ttk.Label(self.medicion_tab, text="SKU:").grid(row=0, column=1, padx=10, pady=5, sticky="w")
        self.sku_entry = ttk.Entry(self.medicion_tab, textvariable=self.sku_var, font=('Helvetica', 10), width=22)
        self.sku_entry.grid(row=0, column=2, padx=10, pady=0, ipadx=15)

        self.medir_button = ttk.Button(self.medicion_tab, text="Medir", command=self.enviar_trama)
        self.medir_button.grid(row=1, rowspan=2, column=1, columnspan=2, padx=10, pady=0, sticky="n")
        
        self.send_button = ttk.Button(self.medicion_tab, text="Enviar",  compound="right", command=self.send_data)
        self.send_button.grid(row=2, rowspan=1, column=1, columnspan=2 ,padx=10, pady=0, sticky="n")

        ttk.Label(self.medicion_tab, text="Largo:").grid(row=0, column=3, padx=10, pady=5, sticky="w")
        self.largo_entry = ttk.Entry(self.medicion_tab, textvariable=self.length_var)
        self.largo_entry.grid(row=0, column=4, padx=10, pady=5)

        ttk.Label(self.medicion_tab, text="Ancho:").grid(row=1, column=3, padx=10, pady=5, sticky="w")
        self.ancho_entry = ttk.Entry(self.medicion_tab, textvariable=self.width_var)
        self.ancho_entry.grid(row=1, column=4, padx=10, pady=5)

        ttk.Label(self.medicion_tab, text="Alto:").grid(row=2, column=3, padx=10, pady=5, sticky="w")
        self.alto_entry = ttk.Entry(self.medicion_tab, textvariable=self.height_var)
        self.alto_entry.grid(row=2, column=4, padx=10, pady=5)

        ttk.Label(self.medicion_tab, text="Peso:").grid(row=3, column=3, padx=10, pady=5, sticky="w")
        self.peso_entry = ttk.Entry(self.medicion_tab, textvariable=self.weight_var)
        self.peso_entry.grid(row=3, column=4, padx=10, pady=5)

        ttk.Label(self.medicion_tab, text="Respuesta:").grid(row=5, column=1, columnspan=2, padx=5, sticky="w")
        self.response_entry = tk.Text(self.medicion_tab, state="disabled", background="#FCFFD0", font=("Arial", 10))
        self.response_entry.config(width=20, height=5)
        self.response_entry.grid(row=6, column=1, columnspan=20, pady=5, sticky="nsew")

        
        # Crear la tabla para mostrar los datos
        columns = ('Sku', 'Largo', 'Ancho', 'Alto', 'Peso', 'Fecha', 'Usuario')
        self.tree = ttk.Treeview(self.medicion_tab, columns=columns, show='headings')

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column('Sku', width=200)
            self.tree.column('Largo', width=50)
            self.tree.column('Ancho', width=50)
            self.tree.column('Alto', width=50)
            self.tree.column('Peso', width=50)
            self.tree.column('Fecha', width=130)
            self.tree.column('Usuario', width=80)
            #self.tree.column(col, width=100)

        self.tree.grid(row=4, column=1, columnspan=20, pady=(10,10))
        
        # Aplicar un estilo con bordes a la tabla
        style = ttk.Style()
        style.configure("Treeview", font=('Helvetica', 9), rowheight=20)
        style.configure("Treeview.Heading", font=('Helvetica', 9))
        style.configure("Treeview.Treeview", borderwidth=1)  # Esto añade bordes alrededor de cada celda
        
        # Crear barras de desplazamiento
        y_scroll = ttk.Scrollbar(self.medicion_tab, orient="vertical", command=self.tree.yview)
        y_scroll.grid(row=4, column=21, sticky='ns')
        self.tree.configure(yscrollcommand=y_scroll.set)
        
        self.medir_button.bind('<Return>', self.on_enter_press)
        self.medir_button.bind('<FocusIn>', self.on_button_focus_in)
        self.medir_button.bind('<FocusOut>', self.on_button_focus_out)
        
        self.send_button.bind('<Return>', self.on_enter_press)
        self.send_button.bind('<FocusIn>', self.on_sendbutton_focus_in)
        self.send_button.bind('<FocusOut>', self.on_sendbutton_focus_out)
        
        self.sku_entry.bind("<Return>", self.cambiar_foco_a_medir)
        
        #Etiquetas del contador
        self.paquetes_enviados_label = tk.Label(self.medicion_tab, text="Envíos exitosos: 0", font=("verdama", 10), fg='green')
        self.paquetes_enviados_label.grid(row=7, column=1, columnspan=2)

        self.paquetes_no_enviados_label = tk.Label(self.medicion_tab, text="Envíos fallidos: 0", font=("Verdana", 10), fg='red')
        self.paquetes_no_enviados_label.grid(row=7,column=3, columnspan=2)

    #Función para funcion del boton de cerrar sesión
    def cerrar_sesion(self):
        self.notebook.tab(0, state="disabled")  # Deshabilitar la pestaña de Medición
        self.notebook.tab(1, state="disabled")  # Deshabilitar la pestaña de Configuración
        self.notebook.tab(2, state="disabled")  # Deshabilitar la pestaña de Configuración
        self.mostrar_ventana_inicio_sesion()  # Mostrar la ventana de inicio de sesión nuevamente

    #CREACIÓN DE COMANDOS PARA FOCUS Y ACCIONES CON ENTER
    # Evento de ENTER en SKU
    def cambiar_foco_a_medir(self, event): 
        self.medir_button.focus_set()
    
    #Evento de enter al tener focus en "Medir" o "Enviar"
    def on_enter_press(self, event):
        if self.is_medirbutton_focused:
            self.enviar_trama()
        elif self.is_sendbutton_focused:
            self.send_data()
            self.sku_entry.focus_set()
    
    #Confirmar cursor en boton "medir" con variable en TRUE
    def on_button_focus_in(self, event): 
        self.is_medirbutton_focused = True 
    
    #Confirmar cursor en boton "medir" con variable en FALSE
    def on_button_focus_out(self, event): 
        self.is_medirbutton_focused = False    
    
    #Confirmar cursor en boton "Enviar" con variable en TRUE
    def on_sendbutton_focus_in(self, event):
        self.is_sendbutton_focused = True
    
    #Confirmar cursor en boton "Enviar" con variable en TRUE
    def on_sendbutton_focus_out(self, event):
        self.is_sendbutton_focused = False


#CREACIÓN DE VENTANA DE CONFIGURACIÓN
    def create_configuracion_tab(self):

        self.url_var = tk.StringVar()
        self.username_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.machine_name_var = tk.StringVar()
        self.ruta_exportacion = tk.StringVar()
        
        # Insertarla en una etiqueta.
        self.colorbackground= "lightgrey"
        self.background = ttk.Label(self.configuracion_tab, background=self.colorbackground)
        self.background.grid(row=0, column=0, rowspan=20,padx=(0,20), sticky="snew")
        
        self.label_montra3 = ttk.Label(self.configuracion_tab, image=self.logo_montra2, background=self.colorbackground)
        self.label_montra3.grid(row=0, column=0, rowspan=3, padx=(10,20), pady=(10,0), sticky="s")
        
        self.label_deprisa1 = ttk.Label(self.configuracion_tab, image=self.logo_deprisa, background=self.colorbackground)
        self.label_deprisa1.grid(row=6, column=0, rowspan=2, padx=(15,20), pady=10, sticky="ew")

        self.label_cubiscan3 = ttk.Label(self.configuracion_tab, image=self.logo_cubiscan2,background=self.colorbackground)
        self.label_cubiscan3.grid(row=3, column=0, rowspan=3, padx=(5,20), sticky="n")
        
        separacion_borde=(0,0)
    
        save_image = customtkinter.CTkImage(Image.open("Icons/save.png").resize((100,100), Image.Resampling.LANCZOS))
        boton_save = customtkinter.CTkButton(self.configuracion_tab, text="Guardar Configuración", corner_radius=1,font=("Helvetica", 14), text_color="#000000", fg_color="#FFFFFF", hover_color="#828890", width=200, height=20, compound="left", image= save_image, command=self.guardar_configuracion)
        boton_save.grid(row=9, column=0, padx=(10,30), pady=10)


        #Botón para crear usuarios
        logout_image = customtkinter.CTkImage(Image.open("Icons/logout.png").resize((100,100), Image.Resampling.LANCZOS))
        boton_logout = customtkinter.CTkButton(self.configuracion_tab, text="Cerrar Sesión", corner_radius=1,font=("Helvetica", 14), text_color="#000000", fg_color="#FFFFFF", hover_color="#828890", width=200, height=20, compound="left", image= logout_image, command=self.cerrar_sesion)
        boton_logout.grid(row=10, column=0, padx=(10,30), pady=(5,5))
        
        ttk.Label(self.configuracion_tab, text=self.texto_licencia ,background=self.colorbackground, font=("Arial", 9)).grid(row=12, rowspan=1, column=0, pady=(5,34), padx=(5,20), sticky="w")
        ttk.Label(self.configuracion_tab, text="DATOS WEB SERVICE:",font=("Helvetica", 13)).grid(row=0, column=1, columnspan=2, padx=separacion_borde, pady=(20,5), sticky="w")
        
        ttk.Label(self.configuracion_tab, text="URL del Web Service:").grid(row=1, padx=separacion_borde, column=1, pady=5, sticky="w")
        url_entry = ttk.Entry(self.configuracion_tab, textvariable=self.url_var, show="*", width=27)
        url_entry.grid(row=1, column=2, pady=5, sticky="w")
        
        ttk.Label(self.configuracion_tab, text="Máquina:").grid(row=4, column=1, padx=separacion_borde, pady=5, sticky="w")
        machine_name_entry = ttk.Entry(self.configuracion_tab, textvariable=self.machine_name_var)
        machine_name_entry.grid(row=4, column=2, pady=5, sticky="w")
        
        ttk.Label(self.configuracion_tab, text="Usuario:").grid(row=2, column=1, padx=separacion_borde, pady=5, sticky="w")
        username_entry = ttk.Entry(self.configuracion_tab, textvariable=self.username_var, show="*")
        username_entry.grid(row=2, column=2, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="Contraseña:").grid(row=3, column=1, padx=separacion_borde, pady=5, sticky="w")
        password_entry = ttk.Entry(self.configuracion_tab, textvariable=self.password_var, show="*")
        password_entry.grid(row=3, column=2, pady=5, sticky="w")

        ttk.Label(self.configuracion_tab, text="EXPORTACIÓN DEL ARCHIVO",font=("Helvetica", 13)).grid(row=5, column=1, columnspan=3, padx=separacion_borde, pady=(20,5), sticky="w")
        ttk.Label(self.configuracion_tab, text="Ruta exportación:").grid(row=6, column=1, padx=separacion_borde, pady=5, sticky="w")
        ruta_exportacion_entry = ttk.Entry(self.configuracion_tab, textvariable=self.ruta_exportacion, width=40)
        ruta_exportacion_entry.grid(row=6, column=2, columnspan=2, pady=5, sticky="w")
        
        
        seleccionar_ruta_image = customtkinter.CTkImage(Image.open("Icons/folder.png").resize((100,100), Image.Resampling.LANCZOS))
        seleccionar_carpeta_button = customtkinter.CTkButton(self.configuracion_tab, text="", corner_radius=1,font=("Helvetica", 14), text_color="#000000", fg_color="#FFFFFF", hover_color="#828890", width=20, height=20, compound="left", image= seleccionar_ruta_image, command=self.seleccionar_carpeta)
        seleccionar_carpeta_button.grid(row=6, column=3, columnspan=4, padx=(125,0), pady=5, sticky="w")
        
        ttk.Label(self.configuracion_tab, text="CONFIGURACIÓN DE COMUNICACIÓN:",font=("Helvetica", 13)).grid(row=8, column=1, columnspan=3, padx=separacion_borde, pady=(20,5), sticky="w")
        ttk.Label(self.configuracion_tab, text="Puertos COM disponibles:").grid(row=9,column=1, padx=separacion_borde, pady=5, sticky="w")
        self.puertos_combobox = ttk.Combobox(self.configuracion_tab)
        self.puertos_combobox.grid(row=9, column=2, padx=5, pady=5)

        self.actualizar_puertos_button = ttk.Button(self.configuracion_tab, text="Actualizar puertos", command=self.listar_puertos)
        self.actualizar_puertos_button.grid(row=9, column=3, padx=5, pady=5, sticky="w")

        self.abrir_puerto_button = ttk.Button(self.configuracion_tab, text="Abrir Puerto", command=self.abrir_puerto)
        self.abrir_puerto_button.grid(row=10, column=2, padx=5, pady=5,  sticky="nw")

        self.cerrar_puerto_button = ttk.Button(self.configuracion_tab, text="Cerrar Puerto", command=self.cerrar_puerto)
        self.cerrar_puerto_button.grid(row=10, column=2, padx=5, pady=5, sticky="ne")
        self.cerrar_puerto_button.configure(state="disabled")

    #Configuración de boton para escoger carpeta de exportación
    def seleccionar_carpeta(self):
        folder_selected = filedialog.askdirectory(title="Seleccione una carpeta de destino")
        self.ruta_exportacion.set(folder_selected)

#CONFIGURACIÓN DE PUERTOS
    #Listar los puertos disponibles
    def listar_puertos(self):
        puertos = [puerto.device for puerto in serial.tools.list_ports.comports()]
        self.puertos_combobox['values'] = puertos
    
    #Abrir puerto seleccionado
    def abrir_puerto(self): 
        puerto_seleccionado = self.puertos_combobox.get()
        if puerto_seleccionado!="" :
            try:
                self.puerto_serial = serial.Serial(puerto_seleccionado, baudrate=9600)
                self.puertos_combobox.configure(state="disabled")
                self.abrir_puerto_button.configure(state="disabled")
                self.cerrar_puerto_button.configure(state="enabled")

                # Guardar el último puerto en el archivo config.ini
                self.guardar_configuracion()
                
                self.data_thread = threading.Thread(target=self.leer_datos)
                self.data_thread.start()
            except Exception as e:
                mensaje = f"{puerto_seleccionado}.\n Error: El puerto especificado no existe"
                messagebox.showerror("Error", mensaje)
    
    #Cerrar el puerto
    def cerrar_puerto(self): 
        if hasattr(self, 'puerto_serial') and self.puerto_serial.is_open:
            self.puerto_serial.close()
            self.puerto_serial = None
            self.puertos_combobox.configure(state="readonly")  # Desbloquear el combobox
            self.abrir_puerto_button.configure(state="enabled")
            self.cerrar_puerto_button.configure(state="disabled")

#CONFIGURACIÓN DE ARCHIVO.INI PARA PRECARGAR Y GUARDAR LOS DATOS
    #Precargar configuración en archivo.ini
    def cargar_configuracion(self):
        mac_actual = self.get_mac_address()  # Usa el método de obtener_mac() definido
        if mac_actual in self.direcciones_mac_permitidas:
            config = configparser.ConfigParser()
            config.read('config.ini')
            if 'Configuracion' in config:
                self.url_var.set(config['Configuracion'].get('url', ''))
                self.username_var.set(config['Configuracion'].get('username', ''))
                self.password_var.set(config['Configuracion'].get('password', ''))
                self.machine_name_var.set(config['Configuracion'].get('machine_name', ''))
                self.ruta_exportacion.set(config['Configuracion'].get('ruta_exportacion', ''))
                # Cargar y establecer el último puerto en el combobox
                ultimo_puerto = config['Configuracion'].get('ultimo_puerto', '')
                self.puertos_combobox.set(ultimo_puerto)
                self.contraseña_actual = config['Configuracion'].get('contraseña_adicional', 'MONTRA101') # Obtener la contraseña
                self.abrir_puerto()
        else: 
            #self.cerrar_puerto()
            mensaje = "Este software solo puede ejecutarse en una computadora autorizada."
            messagebox.showerror("Error", mensaje)
            root.destroy()  # Cierra la aplicación  # Llama a la función para mostrar un mensaje de error y cerrar el programa

    #Guardar configuración en archivo.ini
    def guardar_configuracion(self):
        config = configparser.ConfigParser()
        config.read('config.ini')
        # Guardar los valores de configuración
        config['Configuracion']['url'] = self.url_var.get()
        config['Configuracion']['username'] = self.username_var.get()
        config['Configuracion']['password'] = self.password_var.get()
        config['Configuracion']['machine_name'] = self.machine_name_var.get()
        config['Configuracion']['ruta_exportacion'] = self.ruta_exportacion.get()
        # Obtener el último puerto seleccionado del combobox
        ultimo_puerto = self.puertos_combobox.get()
        config['Configuracion']['ultimo_puerto'] = ultimo_puerto
        config['Configuracion']['contraseña_adicional'] = self.contraseña_actual
        with open('config.ini', 'w') as configfile:
            config.write(configfile)

#CONFIGURACIÓN PARA EXPORTACIÓN DE DATOS
    def exportar_excel(self):
        self.ruta_destino = Path(self.ruta_exportacion.get())
        self.fecha_actual = datetime.datetime.now().strftime("%Y%m%d_%H-%M-%S")
        nombre_archivo = f"CubiScan_{self.fecha_actual}.xlsx"
        ruta_completa = self.ruta_destino / nombre_archivo  # Usar pathlib para construir la ruta

    # Verificar si la carpeta de destino existe
        if self.ruta_exportacion.get() =="" or not self.ruta_destino.exists() or not self.ruta_destino.is_dir():
            self.ruta_destino="Export"
            if not os.path.exists(self.ruta_destino):
                os.makedirs(self.ruta_destino)
            ruta_completa = f"{self.ruta_destino}/CubiScan_{self.fecha_actual}.xlsx" # Usar pathlib para construir la ruta

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Medidas"
        
        # Encabezados
        encabezados = ["SKU", "Largo", "Ancho", "Alto", "Peso", "Fecha", "Usuario"]
        for col_num, encabezado in enumerate(encabezados, 1):
            worksheet.cell(row=1, column=col_num, value=encabezado)

        # Datos
        for row_num, item in enumerate(self.tree.get_children(), 2):
            datos_fila = [self.tree.item(item, 'values')[0], self.tree.item(item, 'values')[1],
                        self.tree.item(item, 'values')[2], self.tree.item(item, 'values')[3],
                        self.tree.item(item, 'values')[4], self.tree.item(item, 'values')[5], 
                        self.tree.item(item, 'values')[6]]
            for col_num, valor in enumerate(datos_fila, 1):
                worksheet.cell(row=row_num, column=col_num, value=valor)

        # Guardar el archivo Excel
        self.guardar_configuracion()
        workbook.save(ruta_completa)

    def exportar_log(self):
        
        text_to_export = self.response_entry.get("1.0", "end-1c")
        # Abrir un cuadro de diálogo para seleccionar la carpeta de destino
        folder_selected = "Log"
        if not os.path.exists(folder_selected):
            os.makedirs(folder_selected)

        if folder_selected:
            # Combinar la carpeta seleccionada con el nombre del archivo
            file_path = f"{folder_selected}/Log_{self.fecha_actual}.txt"

            # Escribir el contenido en el archivo TXT
            with open(file_path, "w") as file:
                file.write(text_to_export)

    def exportar_webservice_error(self):
        log_to_export = self.webservice_error.get("1.0", "end-1c")
        fecha_actual = datetime.datetime.now().strftime("%Y%m%d_%H-%M-%S")
        # Abrir un cuadro de diálogo para seleccionar la carpeta de destino
        log_folder_selected = "Log_WebService_Error"

        if not os.path.exists(log_folder_selected):
            os.makedirs(log_folder_selected)
            
        if log_folder_selected:
            # Combinar la carpeta seleccionada con el nombre del archivo
            log_file_path = f"{log_folder_selected}/Log_{fecha_actual}.txt"

            # Escribir el contenido en el archivo TXT
            with open(log_file_path, "w") as file:
                file.write(log_to_export)
                
        #self.webservice_error.delete("1.0", "end")

#CONFIGURACIÓN PARA RECEPCIÓN PEDIR Y RECIBIR DATOS DE CUBISCAN
    #Solicitar medición
    def enviar_trama(self):
        self.iniciar_espera_datos()
        if hasattr(self, 'puerto_serial') and self.puerto_serial.is_open:
            #self.iniciar_espera_datos()  # Iniciar la espera de datos
            trama = b'\x02M\x03\r\n'  # Trama: <STX>M<ETX><CR><LF>
            enviar_thread = threading.Thread(target=self.enviar_trama_thread, args=(trama,))
            enviar_thread.start()
    
    #Metodo independiente para solicitar medición
    def enviar_trama_thread(self, trama):
        try:
            self.puerto_serial.write(trama)
        except Exception as e:
            print("Error al enviar la trama:", e)  

    #Recepción de datos y estructurar en campos:
    def leer_datos(self):
        while hasattr(self, 'puerto_serial') and self.puerto_serial and self.puerto_serial.is_open:
            try:
                dato = self.puerto_serial.readline().decode('utf-8')
                if dato:
                    self.datos_recibidos = True  # Datos recibidos, detener temporizador
                    if "\x02" in dato and "\x03" in dato:
                        trama = dato.split("\x02")[1].split("\x03")[0]
                        valores = trama.split(",")                        
                        for valor in valores:
                            if valor.startswith("L"):
                                largo = valor.split("L")[1]
                                largo = round(float(largo))  # Convertir a número, redondear
                                self.largo_entry.delete(0, tk.END)
                                self.largo_entry.insert(0, str(largo))
                            elif valor.startswith("W"):
                                ancho = valor.split("W")[1]
                                ancho = round(float(ancho))  # Convertir a número, redondear
                                self.ancho_entry.delete(0, tk.END)
                                self.ancho_entry.insert(0, str(ancho))
                            elif valor.startswith("H"):
                                alto = valor.split("H")[1]
                                alto = round(float(alto))  # Convertir a número, redondear
                                self.alto_entry.delete(0, tk.END)
                                self.alto_entry.insert(0, str(alto))
                            elif valor.startswith("K"):
                                peso = valor.split("K")[1]
                                peso = float(peso)
                                self.peso_entry.delete(0, tk.END)
                                self.peso_entry.insert(0, peso)
                    self.verificar_datos_cubiscan()
            except:
                pass
    
    #Confirmación que ningun dato recibido sea cero.
    def verificar_datos_cubiscan(self):
        # Verificar si alguno de los campos está en 0
        if self.sku_var.get() <= '0' or self.length_var.get() <= '0' or self.width_var.get() <= '0' or self.weight_var.get() <= '0' and self.length_var.get()!= "" or self.width_var.get() == "" or self.height_var.get() == "" or self.weight_var.get() == "":
            self.medir_button.focus_set()
            messagebox.showerror("Error", "Los campos SKU, Largo, Ancho y Alto no pueden ser 0 o estar vacíos.")
            return  # No se envía la información si algún campo es 0
        else:
            self.send_button.focus_set()

    # Metodo para iniciar espera de datos
    def iniciar_espera_datos(self):
        self.datos_recibidos = False
        self.timer = threading.Timer(self.tiempo_espera, self.mostrar_error_sin_datos)
        self.timer.start()

    # Método para mostrar un mensaje de error si no se reciben datos
    def mostrar_error_sin_datos(self):
        if not self.datos_recibidos:
            messagebox.showerror("Error", "No hay comunicación con la CubiScan, revise cables y conexiones.")

    # Agrega este método para detener el temporizador
    def detener_espera_datos(self):
        if hasattr(self, 'timer') or self.timer.is_alive():
            self.timer.cancel()
    

#CONFIGURACIÓN ENVÍO JSON

    def verificar_conexion_internet(self):
        try:
            # Intenta hacer una solicitud GET a un sitio web conocido
            response = requests.get("https://www.google.com")
            if response.status_code == 200:
                return True
        except requests.ConnectionError:
            pass
        return False
    #Configuración del envío de estructura JSON
    def send_data(self):
        # Obtener los valores de los campos
        sku = self.sku_var.get()
        largo = self.length_var.get()
        ancho = self.width_var.get()
        alto = self.height_var.get()
        peso = self.weight_var.get()
        fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        Usuario = self.usuario_registrado

        # Guardar datos en la base de datos
        conn = sqlite3.connect('Montradb.db')
        cursor = conn.cursor()
        cursor.execute('INSERT INTO Usuarios (sku, largo, ancho, alto, peso, fecha, Usuario) VALUES (?, ?, ?, ?, ?, ?, ?)', (sku, largo, ancho, alto, peso, fecha, Usuario))
        conn.commit()
        conn.close()
        
        # Construir el JSON con los datos ingresados
        data = {
            "machine_pid": self.machine_name_var.get(),
            "code": self.sku_var.get(),
            "measure_date": fecha,
            "length": self.length_var.get(),
            "width": self.width_var.get(),
            "heigth": self.height_var.get(),
            "weigth": self.weight_var.get(),
            "unit_type": "cm"
        }
        
        # Verificar si alguno de los campos está en 0
        if sku <= '0' or largo <= '0' or ancho <= '0' or alto <= '0' or peso <= '0':
            self.send_button.focus_set()
            messagebox.showerror("Error", "Los campos SKU, Largo, Ancho y Alto no pueden ser 0 o estar vacíos.")
            return  # No se envía la información si algún campo es 0
        elif sku == "" or largo == "" or alto == "" or ancho == "" or peso=="":
            self.send_button.focus_set()
        else:
            # Mostrar datos en la tabla
            self.tree.insert('', 'end', values=(sku, largo, ancho, alto, peso, fecha, Usuario))


        self.response_entry.tag_config('warning', foreground="red")
        self.response_entry.tag_config('ok', foreground="green")
        # Realizar la solicitud POST al WebService
        if self.verificar_conexion_internet():
            url = self.url_var.get()
            headers = {"Content-Type": "application/json"}
            response = requests.post(url, data=json.dumps(data), headers=headers, auth=(self.username_var.get(), self.password_var.get()))
            #contador
            if response.status_code == 200:
                self.paquetes_enviados += 1
                self.response_entry.config(state=tk.NORMAL)  # Habilita la edición temporalmente
                self.response_entry.insert(tk.END, f"SKU={sku}, Respuesta WS: {response.text}\n", 'ok')
                self.response_entry.config(state=tk.DISABLED)  # Habilita la edición temporalmente   
            else:
                self.paquetes_no_enviados += 1
                self.response_entry.config(state=tk.NORMAL)  # Habilita la edición temporalmente
                self.response_entry.insert(tk.END, f"SKU={sku}, Respuesta WS: {response.text}\n", 'warning')
                self.response_entry.config(state=tk.DISABLED)  # Habilita la edición temporalmente
                self.webservice_error= tk.Text()
                data_text= str(data)
                data_text= data_text.replace("'", '"')
                self.webservice_error.insert(tk.END, data_text)
                self.exportar_webservice_error()
        else:
            self.paquetes_no_enviados += 1
            self.response_entry.config(state=tk.NORMAL)  # Habilita la edición temporalmente
            self.response_entry.insert(tk.END, f"SKU={sku}, Respuesta WS: No hay comunicación con el HOST\n", 'warning')
            self.response_entry.config(state=tk.DISABLED)  # Habilita la edición temporalmente
            self.webservice_error= tk.Text()
            data_text= str(data)
            data_text= data_text.replace("'", '"')
            self.webservice_error.insert(tk.END, data_text)
            self.exportar_webservice_error() 
            messagebox.showerror("Error", "No hay comunicación con el host. Verifique su conexión a internet.")
        

        self.response_entry.see(tk.END)  # Desplaza la vista al final del texto
        self.tree.yview_moveto(1.0)  # Desplaza la vista hacia el final de la tabla
        #actualizar contadores
        self.update_contadores()
        self.sku_var.set("")     # Borra el contenido del campo SKU
        self.length_var.set("")  # Borra el contenido del campo Largo
        self.width_var.set("")   # Borra el contenido del campo Ancho
        self.height_var.set("")  # Borra el contenido del campo Alto
        self.weight_var.set("")  # Borra el contenido del campo Peso


    #Conteos exitoso y fallidos de envío
    def update_contadores(self):
        self.paquetes_enviados_label.config(text=f"Envíos exitosos: {self.paquetes_enviados}")
        self.paquetes_no_enviados_label.config(text=f"Envíos fallidos: {self.paquetes_no_enviados}")


if __name__ == "__main__":
    root = tk.Tk()
    root.resizable(False,False)
    app = SerialInterface(root)
    root.mainloop()
#comentario prueba